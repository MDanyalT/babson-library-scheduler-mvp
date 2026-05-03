"""
Excel exporter — v2 workbook layout
====================================

Sheet order (client-friendly first):

  1. Dashboard         — KPI snapshot + student hours summary table
  2. Schedule          — Client-facing: Date / Day / Time / Student / Pref / Locked
  3. Student Summary   — Per-student hours, gap, constraint status
  4. Violations        — Deduplicated, summary block at top, colour-coded severity
  5. Technical Details — Raw fields: email, reason codes, IDs, override reason
  6. Run Info          — All schedule_runs fields + computed metrics

Public entry points
-------------------
::

    buf = export_schedule_xlsx(conn, run_id)          # BytesIO, seek(0)
    buf = export_week_availability_xlsx(conn, wsd)    # BytesIO, seek(0)

All DB access uses ``conn.execute(text(...), params) / dict(row._mapping)``.
"""

from __future__ import annotations

import json as _json
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.engine import Connection

from app.models.db_models import ViolationType


# ---------------------------------------------------------------------------
# Colour palette  (ARGB hex, no leading #)
# ---------------------------------------------------------------------------
_CLR_NAVY          = "FF1F3864"   # dark navy — header bg
_CLR_WHITE         = "FFFFFFFF"
_CLR_BLACK         = "FF000000"
_CLR_PREFERRED     = "FFD6EAF8"   # light blue — preferred assignment
_CLR_AVAILABLE     = "FFFFFDE7"   # pale yellow — available assignment
_CLR_UNASSIGNED    = "FFFCE4EC"   # pale pink — unfilled shift
_CLR_HARD_SHIFT    = "FFFFF3CD"   # pale amber — hard shift background
_CLR_HARD_VIO      = "FFFFCCCC"   # light red — hard violation
_CLR_SOFT_VIO      = "FFFFF3CD"   # pale amber — soft violation
_CLR_OK            = "FFE8F5E9"   # light green — ok / at-target
_CLR_OVER          = "FFFFCCBC"   # pale orange — over maximum
_CLR_UNDER         = "FFFCE4EC"   # pale pink — under minimum
_CLR_ALT_ROW       = "FFF7F9FC"   # very subtle blue-grey alternating row
_CLR_SECTION_HDR   = "FF2E4A7A"   # section sub-header (slightly lighter navy)
_CLR_KPI_LABEL     = "FFEAF0FB"   # KPI label background

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


# ---------------------------------------------------------------------------
# Public entry point — schedule workbook
# ---------------------------------------------------------------------------

def export_schedule_xlsx(conn: Connection, run_id: str) -> BytesIO:
    """
    Build a formatted Excel workbook for *run_id* and return a BytesIO buffer
    positioned at byte 0 (ready to stream).

    Raises ``KeyError`` if *run_id* does not exist.
    """
    run         = _fetch_run(conn, run_id)
    assignments = _fetch_assignments(conn, run_id)
    summaries   = _fetch_student_summaries(conn, run_id)
    violations  = _fetch_violations(conn, run_id)
    metrics     = _compute_run_metrics(assignments, violations, summaries)

    wb = Workbook()

    _build_dashboard_sheet(wb, run, metrics, summaries)
    _build_schedule_grid_sheet(wb, assignments, run)
    _build_schedule_sheet(wb, assignments)          # now "Schedule List"
    _build_summary_sheet(wb, summaries)
    _build_violations_sheet(wb, violations)
    _build_technical_details_sheet(wb, assignments)
    _build_run_info_sheet(wb, run, metrics)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Public entry point — availability matrix
# ---------------------------------------------------------------------------

def export_week_availability_xlsx(conn: Connection, week_start_date: str) -> BytesIO:
    """
    Export the availability matrix for *week_start_date* as an Excel workbook.
    Rows = students, Columns = shift instances (date + time).
    Cell values: P = preferred, A = available, X = cannot work, blank = no submission.
    """
    students    = _fetch_active_students(conn)
    shifts      = _fetch_shift_instances(conn, week_start_date)
    avail_index = _build_avail_index(conn, week_start_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Availability"

    ws.cell(row=1, column=1, value="Student")
    ws.cell(row=1, column=2, value="Email")
    _hdr(ws.cell(row=1, column=1))
    _hdr(ws.cell(row=1, column=2))

    for col_idx, sh in enumerate(shifts, start=3):
        label = f"{sh['date']}\n{sh['start_time']}–{sh['end_time']}\n[{sh['slot_index']}]"
        cell = ws.cell(row=1, column=col_idx, value=label)
        _hdr(cell)
        _align(cell, wrap=True, h="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28
    ws.row_dimensions[1].height = 48

    for row_idx, st in enumerate(students, start=2):
        s_id = st["id"]
        bg = _CLR_ALT_ROW if row_idx % 2 == 0 else _CLR_WHITE
        _fill_cell(ws.cell(row=row_idx, column=1, value=st["name"]), bg)
        _fill_cell(ws.cell(row=row_idx, column=2, value=st["email"]), bg)

        for col_idx, sh in enumerate(shifts, start=3):
            level = avail_index.get((s_id, sh["id"]))
            if level == "preferred":
                val, cell_bg = "P", _CLR_PREFERRED
            elif level == "available":
                val, cell_bg = "A", _CLR_AVAILABLE
            elif level == "cannot_work":
                val, cell_bg = "X", _CLR_UNASSIGNED
            else:
                val, cell_bg = "", bg
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _fill_cell(cell, cell_bg)
            _align(cell, h="center")

    _freeze_panes(ws, row=2, col=3)
    ws.auto_filter.ref = ws.dimensions

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Sheet 1 — Dashboard
# ---------------------------------------------------------------------------

def _build_dashboard_sheet(
    wb: Workbook,
    run: dict,
    metrics: dict,
    summaries: list[dict],
) -> None:
    ws = wb.create_sheet("Dashboard")

    # ── Section header ──────────────────────────────────────────────────────
    _section_header(ws, 1, 1, "SCHEDULE OVERVIEW", span=4)

    # ── KPI rows ──────────────────────────────────────────────────────────
    mode = run.get("schedule_mode", "weekly")
    mode_label = "Recurring Semester Schedule" if mode == "term_recurring" else "One-Week Schedule"

    kpi_rows: list[tuple] = [
        ("Schedule Type",          mode_label),
    ]
    if mode == "term_recurring":
        kpi_rows.append(("Representative Week Start", run.get("week_start_date", "")))
        kpi_rows.append(("Term Start Date", run.get("term_start_date") or "Not specified"))
        kpi_rows.append(("Term End Date",   run.get("term_end_date")   or "Not specified"))
    else:
        kpi_rows.append(("Week Start Date", run.get("week_start_date", "")))
    kpi_rows += [
        ("Schedule Status",        run.get("status", "").upper()),
        ("Solver Status",          run.get("solver_status", "").upper()),
        ("Generated At",           _fmt_ts(run.get("generated_at", ""))),
        (None, None),   # spacer
        ("Total Shifts",           metrics["total_shifts"]),
        ("Filled Shifts",          metrics["filled"]),
        ("Unfilled Shifts",        metrics["unfilled"]),
        ("Coverage %",             f"{metrics['coverage_pct']}%"),
        (None, None),
        ("Hard Violations",        metrics["hard_count"]),
        ("Soft Violations",        metrics["soft_count"]),
        ("Students Under Minimum", metrics["under_min_count"]),
    ]

    for i, (label, value) in enumerate(kpi_rows, start=2):
        if label is None:
            continue
        lbl_cell = ws.cell(row=i, column=1, value=label)
        val_cell = ws.cell(row=i, column=2, value=value)

        _fill_cell(lbl_cell, _CLR_KPI_LABEL)
        lbl_cell.font = Font(bold=True, color=_CLR_NAVY)

        # Red value if problem
        if label == "Hard Violations" and (metrics["hard_count"] or 0) > 0:
            _fill_cell(val_cell, _CLR_HARD_VIO)
            val_cell.font = Font(bold=True, color="FFCC0000")
        elif label == "Unfilled Shifts" and (metrics["unfilled"] or 0) > 0:
            _fill_cell(val_cell, _CLR_UNASSIGNED)
        elif label == "Students Under Minimum" and (metrics["under_min_count"] or 0) > 0:
            _fill_cell(val_cell, _CLR_UNDER)
        elif label == "Coverage %" and metrics["coverage_pct"] < 100:
            _fill_cell(val_cell, _CLR_SOFT_VIO)
        else:
            _fill_cell(val_cell, _CLR_WHITE)

    # ── Student hours table ────────────────────────────────────────────────
    table_start = 2 + len(kpi_rows) + 1   # one blank row gap
    _section_header(ws, table_start, 1, "STUDENT HOURS SUMMARY", span=5)

    tbl_hdrs = ["Student", "Assigned Hours", "Target Hours", "Gap", "Constraint Status"]
    for col_idx, hdr in enumerate(tbl_hdrs, start=1):
        _hdr(ws.cell(row=table_start + 1, column=col_idx, value=hdr))

    for row_off, sm in enumerate(summaries, start=table_start + 2):
        status = sm.get("constraint_status", "ok")
        bg = {
            "under_minimum": _CLR_UNDER,
            "over_maximum":  _CLR_OVER,
            "at_target":     _CLR_OK,
            "ok":            (_CLR_ALT_ROW if row_off % 2 == 0 else _CLR_WHITE),
        }.get(status, _CLR_WHITE)

        gap = sm.get("hours_vs_target", 0)
        gap_str = f"{gap:+.1f}h" if gap else "0.0h"

        row_vals = [
            sm.get("name", ""),
            sm.get("assigned_hours", 0),
            sm.get("target_hours", 0),
            gap_str,
            status.replace("_", " ").title(),
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_off, column=col_idx, value=val)
            _fill_cell(cell, bg)

    # ── Review Needed section ──────────────────────────────────────────────
    review_start = table_start + 2 + len(summaries) + 2
    _section_header(ws, review_start, 1, "REVIEW NEEDED", span=5)

    review_items: list[tuple[str, str, str]] = []  # (icon, label, value)

    under_min_names = metrics.get("under_min_names", [])
    if under_min_names:
        review_items.append((
            "⚠",
            "Students Under Minimum (8 h)",
            ", ".join(under_min_names),
        ))

    bad_ov = metrics.get("bad_overnight_count", 0)
    if bad_ov:
        review_items.append((
            "⚠",
            "Bad Overnight → Opening Sequences",
            str(bad_ov),
        ))

    low_pref = metrics.get("low_pref_fills", 0)
    total_filled = metrics.get("filled", 0)
    if low_pref:
        pct = round(100.0 * low_pref / total_filled, 1) if total_filled else 0
        review_items.append((
            "ℹ",
            "Shifts Filled by Available (not Preferred)",
            f"{low_pref} ({pct}%)",
        ))

    hard_av = metrics.get("hard_avail_fills", 0)
    if hard_av:
        review_items.append((
            "⚠",
            "Hard Shifts Filled by Available (not Preferred)",
            str(hard_av),
        ))

    if not review_items:
        review_items.append(("✓", "No items requiring review", ""))

    _hdr(ws.cell(row=review_start + 1, column=1, value=""))
    _hdr(ws.cell(row=review_start + 1, column=2, value="Item"))
    _hdr(ws.cell(row=review_start + 1, column=3, value="Detail"))

    for row_off, (icon, label, detail) in enumerate(review_items, start=review_start + 2):
        is_warn = icon == "⚠"
        is_ok   = icon == "✓"
        bg = _CLR_HARD_VIO if is_warn else (_CLR_OK if is_ok else _CLR_SOFT_VIO)

        icon_cell   = ws.cell(row=row_off, column=1, value=icon)
        label_cell  = ws.cell(row=row_off, column=2, value=label)
        detail_cell = ws.cell(row=row_off, column=3, value=detail)
        for c in (icon_cell, label_cell, detail_cell):
            _fill_cell(c, bg)
        label_cell.font = Font(bold=True)
        _align(detail_cell, wrap=True)

    ws.merge_cells(
        start_row=review_start + 1, start_column=3,
        end_row=review_start + 1,   end_column=5,
    )
    ws.merge_cells(
        start_row=review_start, start_column=3,
        end_row=review_start,   end_column=5,
    )

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    _freeze_panes(ws, row=2, col=1)


# ---------------------------------------------------------------------------
# Sheet 2 — Schedule Grid  (client-facing weekly calendar)
# ---------------------------------------------------------------------------

def _build_schedule_grid_sheet(
    wb: Workbook,
    assignments: list[dict],
    run: dict,
) -> None:
    """
    One-page weekly calendar grid.
    Rows = shift windows (sorted by start time).
    Columns = days with shifts (Mon → Sun).
    Cell = assigned student name, colour-coded by preference level.

    All staggered windows are displayed as separate rows — they must NOT be
    collapsed or deduplicated.  Overlapping windows across different students
    are the intentional desk-coverage strategy (one student per shift instance;
    multiple overlapping instances produce continuous coverage during handoffs).
    """
    ws = wb.create_sheet("Schedule Grid")

    # Collect unique dates and unique time windows
    dates_info: dict[str, int] = {}            # date_str → day_of_week
    windows: set[tuple[str, str]] = set()      # (start_time, end_time)

    for asn in assignments:
        sh = asn.get("shift", {})
        d  = sh.get("date")
        dw = sh.get("day_of_week", 0)
        s  = sh.get("start_time")
        e  = sh.get("end_time")
        if d and s and e:
            dates_info[d] = dw
            windows.add((s, e))

    sorted_dates   = sorted(dates_info.keys())
    sorted_windows = sorted(windows)

    if not sorted_dates or not sorted_windows:
        ws.cell(row=1, column=1, value="No assignments to display.")
        return

    # Build lookup: (date, start, end) → first slot_index=0 assignment
    # If multiple slots exist, prefer the filled one.
    lookup: dict[tuple[str, str, str], dict] = {}
    for asn in assignments:
        sh   = asn.get("shift", {})
        key  = (sh.get("date"), sh.get("start_time"), sh.get("end_time"))
        slot = sh.get("slot_index", 0)
        if slot != 0:
            continue
        if key not in lookup:
            lookup[key] = asn
        elif asn.get("student") and not lookup[key].get("student"):
            lookup[key] = asn   # prefer filled over unfilled

    # ── Title row ──────────────────────────────────────────────────────────
    week_str = run.get("week_start_date", "")
    mode     = run.get("schedule_mode", "weekly")
    if mode == "term_recurring":
        title_text = f"Standard Weekly Schedule  (representative week: {week_str})"
        if run.get("term_start_date") and run.get("term_end_date"):
            title_text += f"   |   Term: {run['term_start_date']} → {run['term_end_date']}"
    else:
        title_text = f"Week of {week_str}"
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.fill = PatternFill(fill_type="solid", fgColor=_CLR_SECTION_HDR)
    title_cell.font = Font(bold=True, color=_CLR_WHITE, size=12)
    ws.row_dimensions[1].height = 30

    for col_idx, date_str in enumerate(sorted_dates, start=2):
        dow      = dates_info[date_str]
        day_name = _DAY_NAMES[dow] if 0 <= dow <= 6 else str(dow)
        cell     = ws.cell(row=1, column=col_idx, value=f"{day_name}\n{date_str}")
        _hdr(cell)
        _align(cell, wrap=True, h="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    ws.column_dimensions["A"].width = 16

    # ── Data rows — one per time window ────────────────────────────────────
    for row_idx, (start, end) in enumerate(sorted_windows, start=2):
        time_cell = ws.cell(row=row_idx, column=1, value=f"{start}–{end}")
        time_cell.fill = PatternFill(fill_type="solid", fgColor=_CLR_KPI_LABEL)
        time_cell.font = Font(bold=True, color=_CLR_NAVY)
        _align(time_cell, v="center")
        ws.row_dimensions[row_idx].height = 36

        for col_idx, date_str in enumerate(sorted_dates, start=2):
            asn = lookup.get((date_str, start, end))

            if asn is None:
                # No shift at this day+window
                ws.cell(row=row_idx, column=col_idx, value="")
                continue

            st   = asn.get("student")
            pref = asn.get("preference_level_used", "unassigned")
            hard = asn.get("shift", {}).get("is_hard_shift", False)

            if not st:
                bg, name = _CLR_UNASSIGNED, "UNFILLED"
            elif pref == "preferred":
                bg, name = (_CLR_HARD_SHIFT if hard else _CLR_PREFERRED), st["name"]
            elif pref == "available":
                bg, name = (_CLR_HARD_SHIFT if hard else _CLR_AVAILABLE), st["name"]
            else:
                bg, name = _CLR_WHITE, st["name"]

            pref_label = pref.replace("_", " ").title() if pref and pref != "unassigned" else ""
            cell_text  = f"{name}\n{pref_label}" if pref_label and st else name
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_text)
            _fill_cell(cell, bg)
            _align(cell, wrap=True, h="center", v="center")
            if not st:
                cell.font = Font(bold=True, color="FFCC0000")

    _freeze_panes(ws, row=2, col=2)

    # ── Colour legend (below data) ─────────────────────────────────────────
    legend_row = len(sorted_windows) + 3
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
    legend_items = [
        (_CLR_PREFERRED,  "Preferred assignment"),
        (_CLR_AVAILABLE,  "Available assignment"),
        (_CLR_HARD_SHIFT, "Hard shift (late night / early open)"),
        (_CLR_UNASSIGNED, "Unfilled shift"),
    ]
    for i, (colour, label) in enumerate(legend_items):
        swatch = ws.cell(row=legend_row + 1 + i, column=1, value="")
        _fill_cell(swatch, colour)
        ws.cell(row=legend_row + 1 + i, column=2, value=label)


# ---------------------------------------------------------------------------
# Sheet 3 — Schedule List  (flat row-level list, formerly "Schedule")
# ---------------------------------------------------------------------------

def _build_schedule_sheet(wb: Workbook, assignments: list[dict]) -> None:
    ws = wb.create_sheet("Schedule List")

    headers = [
        "Date", "Day", "Start", "End", "Duration (h)",
        "Slot", "Hard Shift",
        "Student Name", "Preference Level",
        "Locked", "Manual Override",
    ]
    _write_header_row(ws, headers)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row_idx, asn in enumerate(assignments, start=2):
        sh   = asn.get("shift", {})
        st   = asn.get("student")
        pref = asn.get("preference_level_used", "unassigned")

        if pref == "preferred":
            row_bg = _CLR_PREFERRED
        elif pref == "available":
            row_bg = _CLR_AVAILABLE
        else:
            row_bg = _CLR_UNASSIGNED

        # Hard shifts get a subtle overlay when filled
        if sh.get("is_hard_shift") and row_bg not in (_CLR_UNASSIGNED,):
            row_bg = _CLR_HARD_SHIFT

        day_num   = sh.get("day_of_week", 0)
        day_label = _DAY_NAMES[day_num] if 0 <= day_num <= 6 else str(day_num)

        row_values = [
            sh.get("date", ""),
            day_label,
            sh.get("start_time", ""),
            sh.get("end_time", ""),
            sh.get("duration_hours", ""),
            sh.get("slot_index", 0),
            "Yes" if sh.get("is_hard_shift") else "No",
            st["name"] if st else "(Unfilled)",
            pref.replace("_", " ").title() if pref else "",
            "Yes" if asn.get("is_locked") else "No",
            "Yes" if asn.get("is_manual_override") else "No",
        ]

        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _fill_cell(cell, row_bg)
            if col_idx == 8 and not st:   # "(Unfilled)" in bold red
                cell.font = Font(bold=True, color="FFCC0000")

    _set_col_widths(ws, [12, 12, 9, 9, 13, 6, 12, 22, 16, 8, 16])
    _freeze_panes(ws, row=2, col=1)


# ---------------------------------------------------------------------------
# Sheet 4 — Student Summary
# ---------------------------------------------------------------------------

def _build_summary_sheet(wb: Workbook, summaries: list[dict]) -> None:
    ws = wb.create_sheet("Student Summary")

    headers = [
        "Student",
        "Min Hours", "Target Hours", "Max Hours",
        "Assigned Hours", "Gap vs Target",
        "Shifts Assigned", "Preferred", "Available",
        "Constraint Status",
        "Review Note",
    ]
    _write_header_row(ws, headers)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    status_bg = {
        "ok":            _CLR_OK,
        "at_target":     _CLR_OK,
        "over_maximum":  _CLR_OVER,
        "under_minimum": _CLR_UNDER,
    }

    for row_idx, sm in enumerate(summaries, start=2):
        status = sm.get("constraint_status", "ok")
        bg     = status_bg.get(status, _CLR_ALT_ROW if row_idx % 2 == 0 else _CLR_WHITE)

        gap = sm.get("hours_vs_target", 0)
        note = sm.get("review_note", "")
        row_values = [
            sm.get("name", ""),
            sm.get("min_hours", ""),
            sm.get("target_hours", ""),
            sm.get("max_hours", ""),
            sm.get("assigned_hours", 0),
            f"{gap:+.1f}h" if gap else "0.0h",
            sm.get("shifts_assigned", 0),
            sm.get("preferred_shifts", 0),
            sm.get("available_shifts", 0),
            status.replace("_", " ").title(),
            note,
        ]

        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _fill_cell(cell, bg)
            if col_idx == len(headers) and note:   # review note — allow wrapping
                _align(cell, wrap=True)

    _set_col_widths(ws, [24, 10, 13, 10, 15, 14, 16, 10, 10, 20, 52])
    _freeze_panes(ws, row=2, col=1)


# ---------------------------------------------------------------------------
# Sheet 5 — Violations  (deduped + summary)
# ---------------------------------------------------------------------------

def _build_violations_sheet(wb: Workbook, violations: list[dict]) -> None:
    ws = wb.create_sheet("Violations")

    # ── Deduplicate COVERAGE_GAP that duplicates an UNFILLED_SHIFT ─────────
    unfilled_shift_ids = {
        v.get("shift_instance_id")
        for v in violations
        if v.get("violation_type") == ViolationType.UNFILLED_SHIFT
        and v.get("shift_instance_id")
    }
    deduped = [
        v for v in violations
        if not (
            v.get("violation_type") == ViolationType.COVERAGE_GAP
            and v.get("shift_instance_id") in unfilled_shift_ids
        )
    ]

    # ── Summary block ───────────────────────────────────────────────────────
    from collections import Counter
    type_counts: Counter = Counter()
    for v in deduped:
        type_counts[v.get("violation_type", "UNKNOWN")] += 1

    _section_header(ws, 1, 1, "VIOLATION SUMMARY", span=3)
    _hdr(ws.cell(row=2, column=1, value="Violation Type"))
    _hdr(ws.cell(row=2, column=2, value="Severity"))
    _hdr(ws.cell(row=2, column=3, value="Count"))

    summary_row = 3
    hard_types = ViolationType.HARD
    for vtype, count in sorted(type_counts.items(), key=lambda kv: (kv[0] not in hard_types, kv[0])):
        severity = "Hard" if vtype in hard_types else "Soft"
        bg = _CLR_HARD_VIO if severity == "Hard" else _CLR_SOFT_VIO
        _fill_cell(ws.cell(row=summary_row, column=1, value=vtype), bg)
        _fill_cell(ws.cell(row=summary_row, column=2, value=severity), bg)
        _fill_cell(ws.cell(row=summary_row, column=3, value=count), bg)
        summary_row += 1

    # Blank spacer row
    detail_start = summary_row + 1

    # ── Detail rows ─────────────────────────────────────────────────────────
    detail_hdrs = [
        "Violation Type", "Severity", "Shift Date", "Shift Time",
        "Student", "Description",
    ]
    for col_idx, hdr in enumerate(detail_hdrs, start=1):
        _hdr(ws.cell(row=detail_start, column=col_idx, value=hdr))

    ws.auto_filter.ref = (
        f"A{detail_start}:{get_column_letter(len(detail_hdrs))}{detail_start}"
    )

    for row_off, v in enumerate(deduped, start=detail_start + 1):
        severity = v.get("severity", "soft")
        bg = _CLR_HARD_VIO if severity == "hard" else _CLR_SOFT_VIO

        shift_date = ""
        shift_time = ""
        if v.get("shift"):
            sh = v["shift"]
            shift_date = sh.get("date", "")
            shift_time = f"{sh.get('start_time','')}–{sh.get('end_time','')}"

        student_name = v.get("student", {}).get("name", "") if v.get("student") else ""

        row_values = [
            v.get("violation_type", ""),
            severity.title(),
            shift_date,
            shift_time,
            student_name,
            v.get("description", ""),
        ]
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_off, column=col_idx, value=val)
            _fill_cell(cell, bg)
            if col_idx == len(detail_hdrs):   # description — allow wrapping
                _align(cell, wrap=True)

    _set_col_widths(ws, [30, 9, 13, 20, 22, 60])
    ws.row_dimensions[1].height = 18

    if not deduped:
        ws.cell(row=detail_start + 1, column=1, value="No violations found.")
        _fill_cell(ws.cell(row=detail_start + 1, column=1), _CLR_OK)

    _freeze_panes(ws, row=detail_start + 1, col=1)


# ---------------------------------------------------------------------------
# Sheet 6 — Technical Details  (raw/hidden fields)
# ---------------------------------------------------------------------------

def _build_technical_details_sheet(wb: Workbook, assignments: list[dict]) -> None:
    ws = wb.create_sheet("Technical Details")

    headers = [
        "Assignment ID", "Shift Instance ID",
        "Date", "Day", "Start", "End", "Slot",
        "Student ID", "Student Name", "Email",
        "Preference Level", "Reason Codes",
        "Locked", "Manual Override", "Override Reason",
        "Assigned At",
    ]
    _write_header_row(ws, headers)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row_idx, asn in enumerate(assignments, start=2):
        sh  = asn.get("shift", {})
        st  = asn.get("student")
        bg  = _CLR_ALT_ROW if row_idx % 2 == 0 else _CLR_WHITE

        day_num = sh.get("day_of_week", 0)
        row_values = [
            asn.get("id", ""),
            sh.get("id", ""),
            sh.get("date", ""),
            _DAY_NAMES[day_num] if 0 <= day_num <= 6 else str(day_num),
            sh.get("start_time", ""),
            sh.get("end_time", ""),
            sh.get("slot_index", 0),
            st["id"]    if st else "",
            st["name"]  if st else "(Unfilled)",
            st["email"] if st else "",
            asn.get("preference_level_used", ""),
            ", ".join(asn.get("reason_codes", [])),
            "Yes" if asn.get("is_locked") else "No",
            "Yes" if asn.get("is_manual_override") else "No",
            asn.get("override_reason") or "",
            asn.get("assigned_at", ""),
        ]

        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _fill_cell(cell, bg)

    _set_col_widths(ws, [38, 38, 12, 12, 9, 9, 6, 38, 22, 28, 16, 30, 8, 16, 30, 26])
    _freeze_panes(ws, row=2, col=1)


# ---------------------------------------------------------------------------
# Sheet 7 — Run Info
# ---------------------------------------------------------------------------

def _build_run_info_sheet(wb: Workbook, run: dict, metrics: dict) -> None:
    ws = wb.create_sheet("Run Info")

    run_mode = run.get("schedule_mode", "weekly")
    mode_label = "Recurring Semester Schedule" if run_mode == "term_recurring" else "One-Week Schedule"

    fields: list[tuple[str, Any]] = [
        ("Run ID",               run.get("id", "")),
        ("Schedule Type",        mode_label),
        ("Week Start Date",      run.get("week_start_date", "")),
        ("Term Start Date",      run.get("term_start_date") or "Not specified"),
        ("Term End Date",        run.get("term_end_date")   or "Not specified"),
        ("Status",               run.get("status", "")),
        ("Solver Status",        run.get("solver_status", "")),
        ("Solve Time (ms)",      run.get("solve_time_ms", "") or ""),
        ("Objective Score",      run.get("objective_score", "") or ""),
        ("Generated At",         _fmt_ts(run.get("generated_at", ""))),
        ("Published At",         _fmt_ts(run.get("published_at", "") or "")),
        (None, None),
        ("Total Shifts",         metrics["total_shifts"]),
        ("Filled Shifts",        metrics["filled"]),
        ("Unfilled Shifts",      metrics["unfilled"]),
        ("Coverage %",           f"{metrics['coverage_pct']}%"),
        (None, None),
        ("Hard Violations",      metrics["hard_count"]),
        ("Soft Violations",      metrics["soft_count"]),
        ("Students Under Minimum", metrics["under_min_count"]),
        (None, None),
        ("Notes",                run.get("notes", "") or ""),
        (None, None),
        ("MVP Scope Note",
         "term_recurring produces one standard weekly pattern. "
         "Holidays, exam periods, and one-off absences are a future exception layer."),
    ]

    row_idx = 1
    for label, value in fields:
        if label is None:
            row_idx += 1
            continue
        key_cell = ws.cell(row=row_idx, column=1, value=label)
        val_cell = ws.cell(row=row_idx, column=2, value=value)
        _fill_cell(key_cell, _CLR_NAVY)
        key_cell.font = Font(bold=True, color=_CLR_WHITE)
        _fill_cell(val_cell, _CLR_WHITE)
        row_idx += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 44


# ---------------------------------------------------------------------------
# Metric computation
# ---------------------------------------------------------------------------

def _compute_run_metrics(
    assignments: list[dict],
    violations: list[dict],
    summaries: list[dict],
) -> dict:
    total_shifts  = len(assignments)
    filled        = sum(1 for a in assignments if a.get("student"))
    unfilled      = total_shifts - filled
    coverage_pct  = round(100.0 * filled / total_shifts, 1) if total_shifts else 0.0
    hard_count    = sum(1 for v in violations if v.get("severity") == "hard")
    soft_count    = sum(1 for v in violations if v.get("severity") == "soft")
    under_min     = sum(1 for s in summaries if s.get("constraint_status") == "under_minimum")

    # Review-needed metrics
    low_pref_fills     = sum(
        1 for a in assignments
        if a.get("student") and a.get("preference_level_used") == "available"
    )
    hard_avail_fills   = sum(
        1 for a in assignments
        if a.get("student")
        and a.get("shift", {}).get("is_hard_shift")
        and a.get("preference_level_used") == "available"
    )
    bad_overnight_count = sum(
        1 for v in violations
        if v.get("violation_type") == ViolationType.BAD_SEQUENCE_OVERNIGHT
    )
    under_min_names = [
        s.get("name", "") for s in summaries
        if s.get("constraint_status") == "under_minimum"
    ]

    return {
        "total_shifts":       total_shifts,
        "filled":             filled,
        "unfilled":           unfilled,
        "coverage_pct":       coverage_pct,
        "hard_count":         hard_count,
        "soft_count":         soft_count,
        "under_min_count":    under_min,
        "low_pref_fills":     low_pref_fills,
        "hard_avail_fills":   hard_avail_fills,
        "bad_overnight_count": bad_overnight_count,
        "under_min_names":    under_min_names,
    }


# ---------------------------------------------------------------------------
# Database fetch helpers
# ---------------------------------------------------------------------------

def _fetch_run(conn: Connection, run_id: str) -> dict:
    row = conn.execute(
        text("SELECT * FROM schedule_runs WHERE id = :id"),
        {"id": run_id},
    ).fetchone()
    if not row:
        raise KeyError(f"Schedule run not found: {run_id}")
    return dict(row._mapping)


def _fetch_assignments(conn: Connection, run_id: str) -> list[dict]:
    rows = conn.execute(
        text(
            "SELECT "
            "  a.id, a.shift_instance_id, a.student_id, "
            "  a.preference_level_used, a.reason_codes, "
            "  a.is_locked, a.is_manual_override, a.override_reason, a.assigned_at, "
            "  si.date AS shift_date, si.day_of_week, si.start_time, si.end_time, "
            "  si.duration_hours, si.is_hard_shift, si.slot_index, "
            "  s.name AS student_name, s.email AS student_email, "
            "  s.seniority_date AS student_seniority_date "
            "FROM assignments a "
            "JOIN shift_instances si ON si.id = a.shift_instance_id "
            "LEFT JOIN students s ON s.id = a.student_id "
            "WHERE a.run_id = :run_id "
            "ORDER BY si.date, si.start_time, si.slot_index"
        ),
        {"run_id": run_id},
    ).fetchall()

    result = []
    for row in rows:
        r = dict(row._mapping)
        reason_codes: list[str] = []
        try:
            reason_codes = _json.loads(r.get("reason_codes") or "[]")
        except (ValueError, TypeError):
            pass

        asn: dict = {
            "id": r["id"],
            "shift": {
                "id":             r["shift_instance_id"],
                "date":           r["shift_date"],
                "day_of_week":    r["day_of_week"],
                "start_time":     r["start_time"],
                "end_time":       r["end_time"],
                "duration_hours": r["duration_hours"],
                "is_hard_shift":  bool(r["is_hard_shift"]),
                "slot_index":     r["slot_index"],
            },
            "student":              None,
            "preference_level_used": r["preference_level_used"] or "unassigned",
            "reason_codes":         reason_codes,
            "is_locked":            bool(r["is_locked"]),
            "is_manual_override":   bool(r["is_manual_override"]),
            "override_reason":      r.get("override_reason"),
            "assigned_at":          r["assigned_at"],
        }
        if r.get("student_id"):
            asn["student"] = {
                "id":             r["student_id"],
                "name":           r["student_name"] or "",
                "email":          r["student_email"] or "",
                "seniority_date": r["student_seniority_date"] or "",
            }
        result.append(asn)

    return result


def _fetch_student_summaries(conn: Connection, run_id: str) -> list[dict]:
    student_rows = conn.execute(
        text("SELECT * FROM students WHERE is_active = 1 ORDER BY seniority_date, name")
    ).fetchall()
    students = [dict(r._mapping) for r in student_rows]

    # Fetch week so we can look up availability
    run_row = conn.execute(
        text("SELECT week_start_date FROM schedule_runs WHERE id = :id"),
        {"id": run_id},
    ).fetchone()
    wsd = run_row[0] if run_row else ""

    # Per-student availability totals for the week (positive levels only)
    avail_rows = conn.execute(
        text(
            "SELECT av.student_id, COUNT(*) AS slots, "
            "SUM(COALESCE(si.duration_hours, 2.0)) AS hours "
            "FROM availability av "
            "LEFT JOIN shift_instances si ON si.id = av.shift_instance_id "
            "WHERE av.week_start_date = :wsd AND av.level IN ('preferred', 'available') "
            "GROUP BY av.student_id"
        ),
        {"wsd": wsd},
    ).fetchall()
    avail_data: dict[str, dict] = {
        r[0]: {"slots": r[1], "hours": float(r[2] or 0)}
        for r in avail_rows
    }

    asn_rows = conn.execute(
        text(
            "SELECT a.student_id, a.preference_level_used, si.duration_hours "
            "FROM assignments a "
            "JOIN shift_instances si ON si.id = a.shift_instance_id "
            "WHERE a.run_id = :run_id"
        ),
        {"run_id": run_id},
    ).fetchall()

    hours_map:       dict[str, float] = {}
    pref_map:        dict[str, int]   = {}
    avail_map:       dict[str, int]   = {}
    shift_count_map: dict[str, int]   = {}

    for row in asn_rows:
        r   = dict(row._mapping)
        s_id = r.get("student_id")
        if not s_id:
            continue
        hours_map[s_id]       = hours_map.get(s_id, 0.0) + (r["duration_hours"] or 0.0)
        shift_count_map[s_id] = shift_count_map.get(s_id, 0) + 1
        level = r.get("preference_level_used", "")
        if level == "preferred":
            pref_map[s_id]  = pref_map.get(s_id, 0) + 1
        elif level == "available":
            avail_map[s_id] = avail_map.get(s_id, 0) + 1

    summaries = []
    for st in students:
        s_id          = st["id"]
        assigned_hours = round(hours_map.get(s_id, 0.0), 2)
        min_h          = st.get("min_hours", 8)
        max_h          = st.get("max_hours", 20)
        target_h       = st.get("target_hours", 8)

        if assigned_hours < min_h:
            constraint_status = "under_minimum"
        elif assigned_hours > max_h:
            constraint_status = "over_maximum"
        elif abs(assigned_hours - target_h) < 0.1:
            constraint_status = "at_target"
        else:
            constraint_status = "ok"

        review_note = ""
        if constraint_status == "under_minimum":
            av = avail_data.get(s_id, {})
            av_slots = av.get("slots", 0)
            av_hours = av.get("hours", 0.0)   # used only for "insufficient" branch below
            if av_slots == 0:
                review_note = "No availability submitted for this week — needs manual assignment"
            elif av_hours < min_h:
                review_note = (
                    f"Only {av_hours:.1f}h of availability submitted "
                    f"(minimum is {min_h}h) — student cannot reach minimum"
                )
            else:
                review_note = (
                    f"Submitted availability but assigned below {min_h}h — likely "
                    f"constrained by overlap rules, seniority, target balancing, or "
                    f"stronger fit elsewhere. Review manually."
                )

        summaries.append({
            "student_id":        s_id,
            "name":              st.get("name", ""),
            "email":             st.get("email", ""),
            "seniority_date":    st.get("seniority_date", ""),
            "min_hours":         min_h,
            "target_hours":      target_h,
            "max_hours":         max_h,
            "assigned_hours":    assigned_hours,
            "shifts_assigned":   shift_count_map.get(s_id, 0),
            "preferred_shifts":  pref_map.get(s_id, 0),
            "available_shifts":  avail_map.get(s_id, 0),
            "hours_vs_target":   round(assigned_hours - target_h, 2),
            "constraint_status": constraint_status,
            "review_note":       review_note,
        })

    return summaries


def _fetch_violations(conn: Connection, run_id: str) -> list[dict]:
    rows = conn.execute(
        text(
            "SELECT v.*, "
            "  si.date AS shift_date, si.start_time, si.end_time, "
            "  si.day_of_week, si.duration_hours, si.is_hard_shift, si.slot_index, "
            "  s.name AS student_name, s.email AS student_email, "
            "  s.seniority_date AS student_seniority_date "
            "FROM violations v "
            "LEFT JOIN shift_instances si ON si.id = v.shift_instance_id "
            "LEFT JOIN students s ON s.id = v.student_id "
            "WHERE v.run_id = :run_id "
            "ORDER BY "
            "  CASE v.severity WHEN 'hard' THEN 0 ELSE 1 END, "
            "  v.violation_type"
        ),
        {"run_id": run_id},
    ).fetchall()

    result = []
    for row in rows:
        r     = dict(row._mapping)
        viol: dict = {
            "id":             r["id"],
            "violation_type": r["violation_type"],
            "severity":       r["severity"],
            "source":         r.get("source", ""),
            "description":    r.get("description", ""),
            "shift_instance_id": r.get("shift_instance_id"),
            "shift":          None,
            "student":        None,
        }
        if r.get("shift_date"):
            viol["shift"] = {
                "id":             r.get("shift_instance_id", ""),
                "date":           r["shift_date"],
                "day_of_week":    r.get("day_of_week", 0),
                "start_time":     r.get("start_time", ""),
                "end_time":       r.get("end_time", ""),
                "duration_hours": r.get("duration_hours", 0),
                "is_hard_shift":  bool(r.get("is_hard_shift", 0)),
                "slot_index":     r.get("slot_index", 0),
            }
        if r.get("student_id"):
            viol["student"] = {
                "id":             r["student_id"],
                "name":           r.get("student_name", ""),
                "email":          r.get("student_email", ""),
                "seniority_date": r.get("student_seniority_date", ""),
            }
        result.append(viol)

    return result


def _fetch_active_students(conn: Connection) -> list[dict]:
    rows = conn.execute(
        text("SELECT * FROM students WHERE is_active = 1 ORDER BY seniority_date, name")
    ).fetchall()
    return [dict(r._mapping) for r in rows]


def _fetch_shift_instances(conn: Connection, week_start_date: str) -> list[dict]:
    rows = conn.execute(
        text(
            "SELECT * FROM shift_instances WHERE week_start_date = :wsd "
            "ORDER BY date, start_time, slot_index"
        ),
        {"wsd": week_start_date},
    ).fetchall()
    return [dict(r._mapping) for r in rows]


def _build_avail_index(
    conn: Connection,
    week_start_date: str,
) -> dict[tuple[str, str], str]:
    """``{(student_id, shift_instance_id): level}`` — best level per pair."""
    avail_rows = conn.execute(
        text(
            "SELECT student_id, shift_instance_id, day_of_week, "
            "       start_time, end_time, level "
            "FROM availability WHERE week_start_date = :wsd"
        ),
        {"wsd": week_start_date},
    ).fetchall()
    avails = [dict(r._mapping) for r in avail_rows]

    shift_rows = conn.execute(
        text(
            "SELECT id, day_of_week, start_time, end_time "
            "FROM shift_instances WHERE week_start_date = :wsd"
        ),
        {"wsd": week_start_date},
    ).fetchall()
    shifts = [dict(r._mapping) for r in shift_rows]

    shift_meta: dict[str, tuple[int, int, int]] = {}
    for sh in shifts:
        s_min = _hhmm_to_min(sh["start_time"])
        e_min = _hhmm_to_min(sh["end_time"])
        if e_min <= s_min:
            e_min += 1440
        shift_meta[sh["id"]] = (sh["day_of_week"], s_min, e_min)

    index: dict[tuple[str, str], str] = {}
    _LEVEL_RANK = {"preferred": 2, "available": 1, "cannot_work": 0}

    for av in avails:
        s_id  = av["student_id"]
        level = av["level"]

        if av.get("shift_instance_id"):
            key = (s_id, av["shift_instance_id"])
            if _LEVEL_RANK.get(level, -1) > _LEVEL_RANK.get(index.get(key, ""), -1):
                index[key] = level
        else:
            av_day   = av.get("day_of_week")
            av_start = _hhmm_to_min(av.get("start_time", "00:00"))
            av_end   = _hhmm_to_min(av.get("end_time",   "00:00"))
            if av_end <= av_start:
                av_end += 1440

            for sh in shifts:
                if sh["day_of_week"] != av_day:
                    continue
                _, sh_start, sh_end = shift_meta[sh["id"]]
                if av_start <= sh_start and av_end >= sh_end:
                    key = (s_id, sh["id"])
                    if _LEVEL_RANK.get(level, -1) > _LEVEL_RANK.get(index.get(key, ""), -1):
                        index[key] = level

    return index


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _hdr(cell, text_override: str | None = None) -> None:
    """Apply dark navy header style to a cell."""
    if text_override is not None:
        cell.value = text_override
    cell.fill = PatternFill(fill_type="solid", fgColor=_CLR_NAVY)
    cell.font = Font(bold=True, color=_CLR_WHITE)


def _fill_cell(cell, bg: str = _CLR_WHITE) -> None:
    if bg and bg != _CLR_WHITE:
        cell.fill = PatternFill(fill_type="solid", fgColor=bg)


def _align(cell, wrap: bool = False, h: str = "left", v: str = "center") -> None:
    cell.alignment = Alignment(wrap_text=wrap, horizontal=h, vertical=v)


def _section_header(ws, row: int, col: int, title: str, span: int = 4) -> None:
    """Write a full-width section title in the section-header colour."""
    cell = ws.cell(row=row, column=col, value=title)
    cell.fill = PatternFill(fill_type="solid", fgColor=_CLR_SECTION_HDR)
    cell.font = Font(bold=True, color=_CLR_WHITE, size=11)
    ws.row_dimensions[row].height = 20
    # Merge across span columns for a banner look
    if span > 1:
        from openpyxl.utils import get_column_letter
        end_col = get_column_letter(col + span - 1)
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=col + span - 1,
        )


def _write_header_row(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        _hdr(ws.cell(row=1, column=col_idx, value=header))


def _set_col_widths(ws, widths: list[int]) -> None:
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def _freeze_panes(ws, row: int = 2, col: int = 1) -> None:
    ws.freeze_panes = ws.cell(row=row, column=col)


def _hhmm_to_min(t: str) -> int:
    """'HH:MM' → total minutes from midnight. Returns 0 on parse error."""
    try:
        h, m = map(int, t.split(":"))
        return h * 60 + m
    except (ValueError, AttributeError):
        return 0


def _fmt_ts(ts: str) -> str:
    """Trim microseconds and timezone from an ISO timestamp for display."""
    if not ts:
        return ""
    return str(ts)[:19].replace("T", " ")


# ---------------------------------------------------------------------------
# Compatibility aliases
# ---------------------------------------------------------------------------

def build_workbook(conn: Connection, run_id: str):
    """Alias: returns an openpyxl Workbook for the given run."""
    from openpyxl import load_workbook
    buf = export_schedule_xlsx(conn, run_id)
    buf.seek(0)
    return load_workbook(buf)


def workbook_to_bytes(wb) -> bytes:
    """Alias: serialize an openpyxl Workbook to bytes."""
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
